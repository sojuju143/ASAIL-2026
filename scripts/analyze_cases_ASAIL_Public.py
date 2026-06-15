"""
analyze_cases_ASAIL_Public.py
==============================
Public replication script for:

  Yang, K. (2026). "Judicial Writing in Comparative Perspective: An Empirical
  Analysis of Readability and Citation Patterns Across Apex Courts in the
  United Kingdom, Australia, and Singapore."
  Proceedings of the Eighth International Workshop on Automated Semantic
  Analysis of Information in Legal Text (ASAIL 2026), ICAIL 2026.

PURPOSE
-------
Reads a folder of cleaned judicial-text files and outputs an Excel workbook
with per-case readability metrics and citation counts, matching the
methodology described in the paper (§3.2–§3.3).

CORPUS FILE FORMAT
------------------
Each .txt file uses section delimiter lines:

    -----
    HEADNOTES
    -----
    ... headnote text ...
    -----
    CORE JUDGMENT
    -----
    ... judgment text ...
    -----
    FOOTNOTES
    -----
    ... footnote text ...

Readability metrics are computed on CORE JUDGMENT text only (citations
stripped before metric calculation — see §3.2). Citation and academic
reference counts are taken from CORE JUDGMENT + FOOTNOTES combined.
HEADNOTES are excluded throughout (they do not represent judicial writing
style).

METHODOLOGY (matches paper pipeline)
-------------------------------------
Readability (§3.2):
  1. Paragraph-number stripping: leading "45. " tokens removed from line
     starts before sentence tokenisation.
  2. Citation stripping: legal citations removed before FKGL computation
     so that citation abbreviations do not distort syllable/word counts.
     Neutral citations ([2020] UKSC 1), law report citations ([2015] 1 AC
     523), round-bracket citations ((2005) 224 CLR 1), and US volume-first
     citations (285 US 262) are all removed.
  3. Sentence tokenisation: NLTK Punkt trained with ~70 legal abbreviations
     (v., para., s., J., CJ., Ltd., etc.) to suppress false sentence breaks.
     Multi-period abbreviations (e.g., i.e., op. cit.) temporarily replaced
     with soft-hyphen placeholders before tokenisation.
  4. Syllable counting: CMU Pronouncing Dictionary (primary); vowel-group
     heuristic fallback for words not in the dictionary.
  5. FKGL formula: 0.39 × ASL + 11.8 × ASW − 15.59
     where ASL = words/sentences, ASW = syllables/words.
  6. FKRE formula: 206.835 − 1.015 × ASL − 84.6 × ASW
  7. SMOG formula: 3 + √(polysyllabic_words × 30 / sentences)

Citation counting (§3.3):
  - Bracket citations ([YYYY] REPORTER NNN), round-bracket citations
    ((YYYY) vol REPORTER NNN), neutral citations (YYYY COURT NNN), and
    US volume-first citations matched by a multi-pass regex pipeline.
  - Span-based deduplication: overlapping matches counted only once.
  - Jurisdictions: UK, AU, SG, NZ, CAN, USA, EU, HK, MY, IND, OTHER.
  - SCC/SCR disambiguated: bracket form → CAN, round form → IND.

USAGE
-----
  python analyze_cases_ASAIL_Public.py --input /path/to/corpus --output results.xlsx
  python analyze_cases_ASAIL_Public.py --input /path/to/corpus --output results.xlsx --court HCA

REQUIREMENTS
------------
  pip install nltk pandas openpyxl
  python -c "import nltk; nltk.download('punkt'); nltk.download('punkt_tab'); nltk.download('cmudict')"
"""

import os
import re
import sys
import math
import argparse
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import nltk
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================================
# SECTION EXTRACTION
# ============================================================================

def extract_sections(content: str) -> Tuple[str, str, str]:
    """Extract headnotes, core judgment, and footnotes from a cleaned .txt file.

    Handles two formats:
      - Dash-delimited: '-----\\nHEADNOTES\\n-----' (primary corpus format)
      - Plain labels without dashes (fallback for USSC-style files)
    """
    headnotes = ""
    core      = ""
    footnotes = ""

    hn_m = re.search(
        r'-{5,}\s*\n\s*HEADNOTES\s*\n\s*-{5,}\s*\n(.*?)(?=\n\s*-{5,}\s*\n\s*CORE JUDGMENT)',
        content, re.DOTALL | re.IGNORECASE)
    if hn_m:
        headnotes = hn_m.group(1).strip()

    cj_m = re.search(
        r'-{5,}\s*\n\s*CORE JUDGMENT\s*\n\s*-{5,}\s*\n(.*?)(?=\n\s*-{5,}\s*\n\s*FOOTNOTES|\Z)',
        content, re.DOTALL | re.IGNORECASE)
    if cj_m:
        core = cj_m.group(1).strip()

    fn_m = re.search(
        r'-{5,}\s*\n\s*FOOTNOTES\s*\n\s*-{5,}\s*\n(.*)',
        content, re.DOTALL | re.IGNORECASE)
    if fn_m:
        footnotes = fn_m.group(1).strip()

    # Strip SOJU provenance tag from end of footnotes if present
    soju_m = re.search(r'^SOJU\s*-\s*\S', footnotes, re.MULTILINE)
    if soju_m:
        footnotes = footnotes[:soju_m.start()].strip()

    # Fallback: plain label format (no dashes)
    if not core:
        cj_fb = re.search(
            r'(?:^|\n)\s*CORE JUDGMENT\s*\n(.*?)(?=\n\s*FOOTNOTES|\Z)',
            content, re.DOTALL | re.IGNORECASE)
        if cj_fb:
            core = cj_fb.group(1).strip()

        fn_fb = re.search(
            r'(?:^|\n)\s*FOOTNOTES\s*\n(.*)',
            content, re.DOTALL | re.IGNORECASE)
        if fn_fb:
            footnotes = fn_fb.group(1).strip()

        hn_fb = re.search(
            r'(?:^|\n)\s*HEADNOTES\s*\n(.*?)(?=\n\s*CORE JUDGMENT)',
            content, re.DOTALL | re.IGNORECASE)
        if hn_fb:
            headnotes = hn_fb.group(1).strip()

    return headnotes, core, footnotes


# ============================================================================
# METADATA EXTRACTION
# ============================================================================

_NEUTRAL_COURTS_PAT = (
    r'(?:SGCA|SGHC|SGHCF|SGDC|SGMC|UKSC|UKHL|UKPC|'
    r'EWCA(?:\s+(?:Civ|Crim))?|EWHC|'
    r'HCA|FCAFC|FCA|NSWCA|NSWSC|VSC|VSCA|QCA|QSC|'
    r'WASCA|WASC|SASC|NZSC|NZCA|NZHC)'
)

_CASE_HDR  = re.compile(r'^CASE:\s*(.+)',     re.MULTILINE)
_COURT_HDR = re.compile(r'^COURT:\s*(.+)',    re.MULTILINE)
_DATE_HDR  = re.compile(r'^(?:Decision\s+Date|DATE|Date)\s*:\s*(.+?)(?:\n|$)', re.IGNORECASE)


def extract_metadata(content: str, filename: str) -> Dict:
    """Extract case name, citation, year, court, and date from file."""
    fn_base = Path(filename).stem
    header  = content[:600]

    case_m  = _CASE_HDR.search(header)
    court_m = _COURT_HDR.search(header)
    date_m  = _DATE_HDR.search(content)

    case_name = case_m.group(1).strip()  if case_m  else fn_base
    court     = court_m.group(1).strip() if court_m else ''
    date      = date_m.group(1).strip()  if date_m  else ''

    # Citation: scan case name then filename for neutral citation
    cite_pat = re.compile(rf'(\[\d{{4}}\]\s*{_NEUTRAL_COURTS_PAT}\s*\d+)')
    cm = cite_pat.search(case_name) or cite_pat.search(fn_base)
    citation = cm.group(1).strip() if cm else ''

    # Year from citation or filename bracket year
    ym = re.search(r'\[(\d{4})\]', citation or fn_base)
    year = ym.group(1) if ym else ''

    # Fallback date: last full date in headnotes
    if not date:
        months = r'(?:January|February|March|April|May|June|July|August|September|October|November|December)'
        dates = re.findall(rf'(\d{{1,2}}\s+{months}\s+\d{{4}})', content[:2000], re.IGNORECASE)
        if dates:
            date = dates[-1]

    # Clean citation from case name
    clean_name = re.sub(rf'\s*\[\d{{4}}\]\s*{_NEUTRAL_COURTS_PAT}.*$', '', case_name).strip()

    return {
        'case':     clean_name or fn_base,
        'citation': citation,
        'year':     year,
        'court':    court,
        'date':     date,
    }


# ============================================================================
# LEGAL-AWARE SENTENCE TOKENISER
# ============================================================================

# ~70 abbreviations that should NOT trigger sentence boundaries in Punkt.
# Punkt expects lowercase without trailing period.
LEGAL_ABBREVIATIONS = {
    # Versus / case citations
    "v",
    # Paragraph / section / statutory references
    "no", "nos", "para", "paras", "s", "ss", "r", "rr",
    "art", "arts", "ch", "cl", "cls", "sch", "pt", "reg", "regs",
    "p", "pp", "fn", "ed", "eds", "vol", "app",
    # Latin / academic
    "eg", "ie", "cf", "al", "ibid", "id",
    "op", "cit", "loc", "et", "seq",
    "supra", "infra", "ante", "post", "viz", "etc",
    # Company suffixes
    "pte", "sdn", "bhd", "pty", "inc", "corp", "ltd", "co", "plc", "llc", "llp",
    # Personal titles
    "dr", "mr", "mrs", "ms", "prof", "rev", "hon", "rt", "jr", "sr",
    "gen", "col", "sgt",
    # Judicial titles
    "j", "cj", "jj", "ja", "aj", "jca", "cja", "acj", "q", "qc", "kc", "sc", "sjc",
    # Party abbreviations
    "ors", "anor", "anr", "resp", "applt", "respt",
    # Legal-procedural
    "ex", "re", "vs", "div", "dept", "assn", "comm", "dist", "crim",
    # Miscellaneous
    "n", "c", "d", "f", "fig", "govt", "min",
    "approx", "est", "misc", "ref", "acc", "sect",
}

# Multi-period abbreviations: temporarily swapped with soft-hyphen placeholders
# so Punkt does not split at the internal period (e.g. "e.g." → "e\u00ADg\u00AD").
_MULTI_PERIOD = [
    (re.compile(r"\be\.g\.", re.IGNORECASE), "e\u00ADg\u00AD"),
    (re.compile(r"\bi\.e\.", re.IGNORECASE), "i\u00ADe\u00AD"),
    (re.compile(r"\bop\.\s*cit\.", re.IGNORECASE), "op\u00ADcit\u00AD"),
    (re.compile(r"\bloc\.\s*cit\.", re.IGNORECASE), "loc\u00ADcit\u00AD"),
    (re.compile(r"\bet\s+al\.", re.IGNORECASE), "et\u00ADal\u00AD"),
    (re.compile(r"\bet\s+seq\.", re.IGNORECASE), "et\u00ADseq\u00AD"),
]
_RESTORE_MAP = [
    ("e\u00ADg\u00AD", "e.g."), ("i\u00ADe\u00AD", "i.e."),
    ("op\u00ADcit\u00AD", "op. cit."), ("loc\u00ADcit\u00AD", "loc. cit."),
    ("et\u00ADal\u00AD", "et al."), ("et\u00ADseq\u00AD", "et seq."),
]

_legal_tokenizer = None


def _ensure_nltk():
    for pkg in ["punkt", "punkt_tab"]:
        try:
            nltk.data.find(f"tokenizers/{pkg}")
        except LookupError:
            nltk.download(pkg, quiet=True)


def get_legal_tokenizer():
    """Return a Punkt sentence tokeniser customised for legal text."""
    global _legal_tokenizer
    if _legal_tokenizer is not None:
        return _legal_tokenizer
    _ensure_nltk()
    tok = nltk.data.load("tokenizers/punkt_tab/english.pickle")
    tok._params.abbrev_types.update(LEGAL_ABBREVIATIONS)
    _legal_tokenizer = tok
    return _legal_tokenizer


def legal_sent_tokenize(text: str) -> List[str]:
    """Tokenise text into sentences using the legal-aware Punkt tokeniser."""
    t = text
    for pat, repl in _MULTI_PERIOD:
        t = pat.sub(repl, t)
    sents = get_legal_tokenizer().tokenize(t)
    restored = []
    for s in sents:
        for placeholder, original in _RESTORE_MAP:
            s = s.replace(placeholder, original)
        restored.append(s)
    return restored


# ============================================================================
# CITATION STRIPPING FOR READABILITY METRICS
# ============================================================================
# Citations are stripped from core text before FKGL/FKRE/SMOG calculation.
# This prevents citation abbreviations (e.g. "CLR", "WLR") from being
# counted as words and distorting syllable/word ratios.
# Citations are NOT stripped when counting citation frequencies (§3.3).

_LEADING_PARA_NUM = re.compile(r"(?m)^\s*\d{1,3}\.\s+")

# Neutral citations: [2020] UKSC 1, [2009] HCA 14, 2005 CanLII 12345
_STRIP_NEUTRAL = re.compile(
    r'\[\s*\d{4}\s*\]\s+'
    r'(?:UKSC|UKHL|UKPC|EWCA\s+(?:Civ|Crim)|EWHC|UKUT|'
    r'CSOH|CSIH|SGCA|SGHC|SGHCF|SGDC|SGMC|'
    r'HCA|FCAFC|FCA|NSWCA|NSWSC|VSC|VSCA|QCA|QSC|'
    r'WASCA|WASC|SASC|SASFC|TASSC|ACTSC|ACTCA|'
    r'NZSC|NZCA|NZHC|'
    r'HKCFA|HKCA|HKCFI|'
    r'ABCA|ABQB|ONCA|ONSC|BCCA|BCSC)'
    r'\s+\d+'
)
_STRIP_CANLII = re.compile(r'\b\d{4}\s+CanLII\s+\d+')

# Law report citations with volume: [2020] 1 AC 123, [2015] 2 SLR(R) 456
_STRIP_REPORT = re.compile(
    r'\[\s*\d{4}\s*\]\s+\d+\s+'
    r'(?:AC|WLR|QB|KB|Ch|Fam|All\s+ER|SLR(?:\(R\))?|MLJ|'
    r'ICR|IRLR|Cr\s+App\s+R(?:\s+\(S\))?|BCLC|BCC|FSR|RPC|STC|WTLR|'
    r'NZLR|DLR(?:\s+\(\d+[a-z]+\))?|SCR|OR(?:\s+\(\d+[a-z]+\))?|'
    r'Bus\s+LR|Bank\s+LR|Imm\s+AR|'
    r'Lloyd(?:\'?s)?\s+Rep(?:\s+(?:PN|IR))?|Ll\s+L\s+Rep|'
    r'P\s+&\s+CR|ACLR|ACSR|ACLC|EHRR|CMLR|BHRC'
    r')\s+\d+'
)

# Round-bracket citations: (2005) 224 CLR 123
_STRIP_ROUND = re.compile(
    r'\(\d{4}\)\s+\d+\s+'
    r'(?:CLR|ALR|ALJR|FLR|NSWLR|VR|SASR|WAR|'
    r'US|S\s+Ct|L\s+Ed(?:\s+2d)?|'
    r'F\.?\s*3d|F\.?\s*2d|F\.?\s*Supp(?:\.?\s*2d)?|'
    r'DLR(?:\s+\(\d+[a-z]+\))?|OR(?:\s+\(\d+[a-z]+\))?|'
    r'App\s+Cas|Ch\s+D|Ch\s+App|ER|Beav|Ves\s+Jun|HLC|HL\s+Cas|'
    r'Cr\s+App\s+R(?:\s+\(S\))?|Ll\s+L\s+Rep|P\s+&\s+CR|'
    r'ACLR|ACSR|ACLC|A\s+Crim\s+R|NS\s+WLR'
    r')\s+\d+'
)

# US volume-first: 285 US 262, 123 F.3d 456
_STRIP_US = re.compile(
    r'\b\d{1,3}\s+'
    r'(?:U\.?\s*S\.?|S\.?\s*Ct\.?|L\.?\s*Ed\.?(?:\s*2d)?|'
    r'F\.?\s*3d|F\.?\s*2d|F\.?\s*Supp(?:\.?\s*2d)?)'
    r'\s+\d+'
)

# Broad bracket citation catch-all: [YYYY] REPORTER NNN
_STRIP_BRACKET_BROAD = re.compile(r'\[\s*\d{4}\s*\]\s+[A-Z][A-Z0-9]+\s+\d+\b')

# Pinpoints: "(at [45]–[47])", "at p. 23", "at [8]"
_STRIP_PINPOINT_PAREN = re.compile(
    r'\(at\s+(?:pages?\s+\d+(?:\s*[-\u2013]\s*\d+)?'
    r'|pp?\.\s*\d+(?:\s*[-\u2013]\s*\d+)?'
    r'|\[\d+\](?:\s*[-\u2013,]\s*\[\d+\])*)\)'
)
_STRIP_PINPOINT_PARA = re.compile(r'\bat\s+\[\d+\](?:\s*[-\u2013,]\s*\[\d+\])*')
_STRIP_PINPOINT_PAGE = re.compile(
    r'\bat\s+(?:pages?\s+\d+(?:\s*[-\u2013]\s*\d+)?|pp?\.\s*\d+(?:\s*[-\u2013]\s*\d+)?)'
)


def prepare_text_for_metrics(text: str) -> str:
    """Prepare core judgment text for readability metric calculation.

    Steps (in order):
      1. Remove leading paragraph numbers ('45. ' at line start)
      2. Remove broad bracket citations
      3. Remove neutral, law report, round-bracket, and US citations
      4. Remove citation pinpoints
      5. Normalise whitespace
    """
    t = _LEADING_PARA_NUM.sub('', text)
    t = _STRIP_BRACKET_BROAD.sub('', t)
    t = _STRIP_NEUTRAL.sub('', t)
    t = _STRIP_CANLII.sub('', t)
    t = _STRIP_REPORT.sub('', t)
    t = _STRIP_ROUND.sub('', t)
    t = _STRIP_US.sub('', t)
    t = _STRIP_PINPOINT_PAREN.sub('', t)
    t = _STRIP_PINPOINT_PARA.sub('', t)
    t = _STRIP_PINPOINT_PAGE.sub('', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


# ============================================================================
# SYLLABLE COUNTING  (CMU Pronouncing Dictionary + vowel-group fallback)
# ============================================================================

_vowel_re = re.compile(r'[aeiouAEIOU]+')
_cmudict   = None
_HAS_CMU   = False


def _load_cmudict():
    global _cmudict, _HAS_CMU
    if _cmudict is not None:
        return
    try:
        nltk.data.find("corpora/cmudict")
    except LookupError:
        nltk.download("cmudict", quiet=True)
    try:
        _cmudict = nltk.corpus.cmudict.dict()
        _HAS_CMU = True
    except Exception:
        _cmudict = {}
        _HAS_CMU = False


def count_syllables(word: str) -> int:
    """Count syllables using CMU dict (primary) or vowel-group heuristic."""
    _load_cmudict()
    clean = re.sub(r"[^a-z]", "", word.lower())
    if not clean:
        return 0
    if _HAS_CMU and clean in _cmudict:
        return max(1, sum(1 for ph in _cmudict[clean][0] if ph[-1].isdigit()))
    return max(1, len(_vowel_re.findall(clean)))


# ============================================================================
# READABILITY METRICS
# ============================================================================

def word_count(text: str) -> int:
    return len(re.findall(r'\b\w+\b', text)) if text.strip() else 0


def compute_readability(text: str) -> Tuple:
    """Compute FKGL, FKRE, SMOG, and average sentence length.

    Input should be pre-processed with prepare_text_for_metrics().
    Returns (fkgl, fkre, smog, avg_sent_length) or ('','','','') on failure.
    """
    if not text or len(text.split()) < 30:
        return '', '', '', ''
    try:
        # Sentence count via legal-aware Punkt
        t_tok = text
        for pat, repl in _MULTI_PERIOD:
            t_tok = pat.sub(repl, t_tok)
        sents = [s for s in get_legal_tokenizer().tokenize(t_tok) if len(s.split()) > 1]
        n_sent = len(sents)
        if n_sent < 3:
            return '', '', '', ''

        # Word/syllable counts from text as-is
        words   = re.findall(r'\b\w+\b', text)
        n_words = len(words)
        if n_words == 0:
            return '', '', '', ''
        n_syl  = sum(count_syllables(w) for w in words)
        n_poly = sum(1 for w in words if count_syllables(w) >= 3)

        asl  = n_words / n_sent
        asw  = n_syl   / n_words
        fkgl = round(0.39 * asl + 11.8 * asw - 15.59, 1)
        fkre = round(206.835 - 1.015 * asl - 84.6 * asw, 1)
        smog = round(3 + (n_poly * 30 / n_sent) ** 0.5, 1)
        return fkgl, fkre, smog, round(asl, 1)
    except Exception:
        return '', '', '', ''


# ============================================================================
# CITATION JURISDICTION CLASSIFICATION
# ============================================================================

# Reporter sets by jurisdiction.
# Ordered check: HK → MY → SG before UK to avoid MLJ/BLR substring collisions.

UK_REPORTERS = {
    'UKHL', 'UKSC', 'UKPC', 'AC', 'WLR', 'QB', 'KB', 'Ch', 'Fam',
    'EWCA', 'EWHC', 'All ER', 'ICR', 'IRLR', 'Cr App R', 'Cr App R (S)',
    'BCLC', 'BCC', 'FSR', 'RPC', 'STC', 'WTLR',
    'CSIH', 'CSOH', 'UKUT', 'UKEAT',
    'HLC', 'BLR', 'LGR', 'RTR', 'P & CR', 'BPIR',
    'App Cas', 'Cl & Fin', 'HL Cas',
    'Cowp', 'Ch D', 'Ch App', 'ER', 'Beav', 'LJ Ch', 'Morr', 'Bing', 'QBD', 'Ex',
    'Bus LR', 'Bank LR', 'Imm AR',
    'Lloyd Rep', "Lloyd's Rep", 'Ll L Rep',
    'Lloyd Rep PN', "Lloyd's Rep PN", 'Lloyd Rep IR', "Lloyd's Rep IR",
    'TLR', 'LR', 'Hare',
}

AU_REPORTERS = {
    'HCA', 'FCAFC', 'FCA', 'NSWCA', 'NSWSC', 'VSC', 'VSCA', 'QCA', 'QSC',
    'WASCA', 'WASC', 'SASC', 'SASFC', 'TASSC', 'ACTSC', 'ACTCA',
    'CLR', 'ALR', 'ALJR', 'FLR', 'NSWLR', 'VR', 'SASR', 'WAR', 'Qd R',
    'VLR', 'NTLR', 'ACTR',
    'ACLR', 'ACSR', 'ACLC', 'ABC', 'A Crim R',
    'FCR', 'NSWCCA', 'IPR', 'MVR', 'ATR', 'LGERA', 'ALD',
    'NS WLR',
}

SG_REPORTERS = {
    'SGCA', 'SGHC', 'SGHCF', 'SGDC', 'SGMC', 'SLR', 'SGHCR', 'SGDT', 'BLR',
}

NZ_REPORTERS = {
    'NZSC', 'NZCA', 'NZHC', 'NZLR', 'NZAR', 'NZCLC',
    'NZFLR', 'NZFC', 'FRNZ', 'NZRMA', 'NZTC', 'NZCPR',
    'PRNZ', 'CRNZ', 'TCLR', 'ERNZ', 'NZCCLR', 'HRNZ', 'NZIPT', 'NZELR',
    'NZDC', 'NZMED', 'NZPB', 'NZRRD', 'NZGCR',
}

CAN_REPORTERS = {
    'FC', 'FCA', 'ONCA', 'ONSC', 'BCCA', 'BCSC', 'ABCA', 'ABQB',
    'DLR', 'OR', 'AR', 'BCLR', 'WWR', 'RFL', 'CCC',
    'OAC', 'Can Bus LJ', 'CanLII',
}

USA_REPORTERS = {
    'US', 'S Ct', 'L Ed', 'L Ed 2d',
    'F', 'F 2d', 'F 3d', 'F Supp', 'F Supp 2d',
    'So', 'So 2d', 'NE', 'NE 2d', 'NW', 'NW 2d', 'SE', 'SE 2d', 'SW', 'SW 2d',
    'P 2d', 'P 3d', 'A 2d', 'A 3d',
    'F.3d', 'F.2d', 'F.Supp', 'F.Supp.2d',
    'N.E.', 'N.E.2d', 'N.W.', 'N.W.2d', 'S.E.', 'S.E.2d', 'S.W.', 'S.W.2d',
    'So.', 'So.2d', 'U.S.', 'S.Ct.', 'L.Ed.', 'L.Ed.2d',
}

EU_REPORTERS  = {'ECR', 'CMLR', 'EUECJ', 'ECHR', 'EHRR', 'DR', 'ETMR', 'BHRC'}
HK_REPORTERS  = {'HKCFA', 'HKCA', 'HKCFI', 'HKC', 'HKLRD', 'HKCU', 'HKLR'}
MY_REPORTERS  = {'MLJ', 'MLJU', 'MLRA', 'MYCA', 'MYHC', 'MYFC', 'CLJ'}
IND_REPORTERS = {'AIR', 'Bom', 'Cal', 'Mad', 'All', 'Del', 'Kar', 'Ker'}

# SCC/SCR appear in both Canadian and Indian law reports —
# disambiguated by citation form: bracket → CAN, round → IND.
AMBIGUOUS_CAN_IND = {'SCC', 'SCR'}

# Academic journal abbreviations that superficially resemble law report codes.
# Matches suppressed so they do not appear in citation counts.
ACADEMIC_JOURNALS = {
    'LQR', 'LMCLQ', 'SJLS', 'OJLS', 'CLP', 'JBL', 'ICLQ',
    'AJCL', 'SAcLJ', 'NZLJ',
}

# Exact-match overrides for reporters prone to substring false positives.
# Checked before the set-membership loop.
_EXACT_OVERRIDE = {
    'P':          'UK',    # Probate Reports (UK) — not US Pacific (P 2d/P 3d)
    'A CRIM R':   'AU',    # Australian Criminal Reports (not UK 'AC')
    'BUS LR':     'UK',
    'BANK LR':    'UK',
    'IMM AR':     'UK',
    'LL L REP':   'UK',
    'LLOYD REP':  'UK',
    "LLOYD'S REP":'UK',
    'CR APP R':   'UK',
    'CR APP R (S)':'UK',
    'P & CR':     'UK',
    'APP CAS':    'UK',
    'CH D':       'UK',
    'CH APP':     'UK',
    'LJ CH':      'UK',
    'VES JUN':    'UK',
    'HL CAS':     'UK',
    'ALL ER':     'UK',
    'US':         'USA',
    'S CT':       'USA',
    'L ED':       'USA',
    'L ED 2D':    'USA',
    'F 2D':       'USA',
    'F 3D':       'USA',
    'F SUPP':     'USA',
    'F SUPP 2D':  'USA',
    'NE':         'USA',
    'NE 2D':      'USA',
    'NW':         'USA',
    'NW 2D':      'USA',
    'SE':         'USA',
    'SE 2D':      'USA',
    'SW':         'USA',
    'SW 2D':      'USA',
    'SO':         'USA',
    'SO 2D':      'USA',
    'CAN BUS LJ': 'CAN',
    'CANLII':     'CAN',
    'OAC':        'CAN',
    'NZCLC':      'NZ',
    'ACLR':       'AU',
    'ACSR':       'AU',
    'ACLC':       'AU',
    'ABC':        'AU',
    'NS WLR':     'AU',
    'LR':         'UK',
    'HARE':       'UK',
}


def classify_reporter(reporter: str, is_bracket_form: bool = True) -> Optional[str]:
    """Map a reporter abbreviation to a jurisdiction key.

    Returns one of: 'UK', 'AU', 'SG', 'NZ', 'CAN', 'USA', 'EU', 'HK', 'MY',
    'IND', 'OTHER', or None (if the abbreviation is an academic journal).

    is_bracket_form: True if citation used '[YYYY]' brackets (used to
    disambiguate SCC/SCR between Canada and India).
    """
    rc  = reporter.strip().upper()
    rc  = re.sub(r'\s+\d+$', '', rc)              # strip trailing digits
    rnd = rc.replace('.', '').replace(' ', '')     # normalised no-dots form

    # Suppress academic journals
    for aj in ACADEMIC_JOURNALS:
        if aj.upper() in (rc, rnd):
            return None

    # Exact overrides first
    if rc  in _EXACT_OVERRIDE: return _EXACT_OVERRIDE[rc]
    if rnd in _EXACT_OVERRIDE: return _EXACT_OVERRIDE[rnd]

    # SCC/SCR: bracket → CAN, round → IND
    rk = re.match(r'[A-Za-z]+', reporter.strip())
    if rk and rk.group().upper() in AMBIGUOUS_CAN_IND:
        return 'CAN' if is_bracket_form else 'IND'

    def _match(rep_set):
        for rep in rep_set:
            ru = rep.upper()
            if ru in rc or rc in ru or ru in rnd or rnd in ru:
                return True
        return False

    # Ordered: HK/MY before SG to avoid MLJ/BLR collisions
    if _match(HK_REPORTERS):  return 'HK'
    if _match(MY_REPORTERS):  return 'MY'
    if _match(SG_REPORTERS):  return 'SG'
    if _match(UK_REPORTERS):  return 'UK'
    if _match(AU_REPORTERS):  return 'AU'
    if _match(USA_REPORTERS): return 'USA'
    if _match(CAN_REPORTERS): return 'CAN'
    if _match(IND_REPORTERS): return 'IND'
    if _match(NZ_REPORTERS):  return 'NZ'
    if _match(EU_REPORTERS):  return 'EU'

    return 'OTHER'


# ============================================================================
# CITATION EXTRACTION & COUNTING
# ============================================================================

# Primary pattern: single-word reporters
# Group layout:
#   bracket form: groups 0,1,2 = year, reporter, number
#   round form:   groups 3,4,5,6 = year, volume, reporter, number
_CITE_PRIMARY = re.compile(
    r'\[(\d{4})\]\s*(?:(\d+)\s+)?([A-Z][A-Za-z]*(?:\s*\([A-Za-z]+\))?)\s*(\d+)'
    r'|'
    r'\((\d{4})\)\s*(\d+)\s+([A-Z][A-Za-z]+(?:\s*\d*[A-Za-z]*)?)\s*(\d+)'
)

# Extended patterns: multi-word reporters, dotted US forms, CanLII
_CITE_EXTENDED = [
    # [YYYY] vol REPORTER(variant) NNN — e.g. [2015] 2 SLR(R) 456
    re.compile(r'\[(\d{4})\]\s+(\d+)\s+([A-Z][A-Za-z]+(?:\([A-Z]+\))?)\s+(\d+)'),
    # [YYYY] vol Multi-word Reporter NNN
    re.compile(
        r'\[(\d{4})\]\s+(\d+)\s+'
        r'((?:All\s+ER|Cr\s+App\s+R(?:\s+\(S\))?|Bus\s+LR|Bank\s+LR|Imm\s+AR|'
        r'Ll\s+L\s+Rep|Lloyd\'?s?\s+Rep(?:\s+(?:PN|IR))?|'
        r'P\s+&\s+CR|App\s+Cas|Ch\s+D|Ch\s+App|LJ\s+Ch|'
        r'A\s+Crim\s+R|Can\s+Bus\s+LJ|Ves\s+Jun|HL\s+Cas|'
        r'NS\s+WLR|S\s+Ct|L\s+Ed))'
        r'\s+(\d+)'
    ),
    # [YYYY] Multi-word Reporter NNN (no volume)
    re.compile(
        r'\[(\d{4})\]\s+'
        r'((?:Bus\s+LR|Bank\s+LR|Imm\s+AR|Ll\s+L\s+Rep|'
        r'Lloyd\'?s?\s+Rep(?:\s+(?:PN|IR))?|P\s+&\s+CR|Cr\s+App\s+R(?:\s+\(S\))?|All\s+ER|P))'
        r'\s+(\d+)'
    ),
    # (YYYY) vol Multi-word Reporter NNN
    re.compile(
        r'\((\d{4})\)\s+(\d+)\s+'
        r'((?:App\s+Cas|Ch\s+D|Ch\s+App|LJ\s+Ch|Ves\s+Jun|HL\s+Cas|'
        r'A\s+Crim\s+R|Cr\s+App\s+R(?:\s+\(S\))?|Ll\s+L\s+Rep|'
        r'Lloyd\'?s?\s+Rep(?:\s+(?:PN|IR))?|P\s+&\s+CR|'
        r'Can\s+Bus\s+LJ|NS\s+WLR|S\s+Ct|L\s+Ed))'
        r'\s+(\d+)'
    ),
    # Dotted US reporters: 123 F.3d 456, 123 N.E.2d 789
    re.compile(
        r'\b(\d{1,3})\s+'
        r'((?:F\.\s*3d|F\.\s*2d|F\.\s*Supp(?:\.\s*2d)?|'
        r'N\.E\.(?:\s*2d)?|N\.W\.(?:\s*2d)?|S\.E\.(?:\s*2d)?|'
        r'S\.W\.(?:\s*2d)?|So\.(?:\s*2d)?|'
        r'P\.\s*2d|P\.\s*3d|A\.\s*2d|A\.\s*3d|'
        r'U\.S\.|S\.\s*Ct\.|L\.\s*Ed\.(?:\s*2d)?))'
        r'\s+(\d+)'
    ),
    # CanLII neutral: 2005 CanLII 12345
    re.compile(r'\b(\d{4})\s+(CanLII)\s+(\d+)'),
    # DLR/OR with series: (1978) 87 DLR (3d) 112
    re.compile(r'\((\d{4})\)\s+(\d+)\s+((?:DLR|OR|OAC)\s*\(\d+[a-z]+\))\s+(\d+)'),
    # US bare volume-first: 285 US 262, 123 U.S. 456
    re.compile(r'\b(\d{1,3})\s+(US)\s+(\d+)'),
    re.compile(r'\b(\d{1,3})\s+(U\.S\.)\s+(\d+)'),
]


def _spans_overlap(a: Tuple[int, int], b: Tuple[int, int]) -> bool:
    return a[0] < b[1] and b[0] < a[1]


def count_citations(text: str) -> Dict[str, int]:
    """Count case citations by jurisdiction from a text block.

    Multi-pass extraction with full span-based overlap deduplication:
      Pass 1 — primary pattern (single-word reporters)
      Pass 2 — extended patterns (multi-word, dotted US, CanLII)

    Returns a dict with keys:
      'UK', 'AU', 'SG', 'NZ', 'CAN', 'USA', 'EU', 'HK', 'MY', 'IND', 'OTHER'
    """
    juri_keys = ['UK', 'AU', 'SG', 'NZ', 'CAN', 'USA', 'EU', 'HK', 'MY', 'IND', 'OTHER']
    counts = {k: 0 for k in juri_keys}
    used_spans: List[Tuple[int, int]] = []

    def _process(span, reporter, is_bracket):
        for existing in used_spans:
            if _spans_overlap(span, existing):
                return
        juri = classify_reporter(reporter, is_bracket_form=is_bracket)
        if juri is None:
            return  # academic journal — skip
        used_spans.append(span)
        counts[juri if juri in counts else 'OTHER'] += 1

    # Pass 1: primary pattern
    for m in _CITE_PRIMARY.finditer(text):
        g = m.groups()
        if g[0]:   # bracket form: [YYYY] [vol] REPORTER NNN
            reporter     = (g[1] or g[2]).strip()
            is_bracket   = True
        else:      # round form: (YYYY) vol REPORTER NNN
            reporter     = g[6].strip() if g[6] else ''
            is_bracket   = False
        if not reporter or reporter.isdigit():
            continue
        reporter_clean = reporter.replace('.', '').strip()
        _process((m.start(), m.end()), reporter_clean, is_bracket)

    # Pass 2: extended patterns
    for pat in _CITE_EXTENDED:
        for m in pat.finditer(text):
            grps = [g for g in m.groups() if g]
            if len(grps) == 3:
                g0, g1, g2 = grps
                if g1 == 'CanLII':
                    reporter, is_bracket = g1, False
                elif g0.isdigit() and len(g0) <= 3:
                    reporter, is_bracket = g1, False
                else:
                    reporter, is_bracket = g1, True
            elif len(grps) == 4:
                reporter   = grps[2]
                is_bracket = '[' in m.group()
            else:
                continue
            reporter_clean = reporter.replace('.', '').strip()
            _process((m.start(), m.end()), reporter_clean, is_bracket)

    return counts


# ============================================================================
# ACADEMIC REFERENCE DETECTION
# ============================================================================

def count_academic_references(text: str) -> int:
    """Count unique academic references (journal articles and books).

    Uses span-based overlap removal so that a single reference matched by
    multiple patterns is counted only once.
    """
    journal_patterns = [
        r'\(\d{4}\)\s+\d+\s+[A-Z][A-Za-z]+(?:\s+[A-Za-z]+)+\s+(?:Law|Legal|Journal|Review|Quarterly|University|Studies)\s+[A-Za-z]*\s*\d*',
        r'\d+\s+(?:Law\s+)?(?:Journal|Review|Quarterly|L\.?\s*J\.?|L\.?\s*Rev\.?|L\.?\s*Q\.?)',
        r'\b(?:LQR|CLJ|OJLS|CLP|Sing\.?\s*L\.?\s*Rev\.?|SJLS)\b',
        r'\b(?:Law\s+Quarterly\s+Review|Cambridge\s+Law\s+Journal|Oxford\s+Journal\s+of\s+Legal\s+Studies)\b',
        r'\b(?:Yale\s+L\.?\s*J\.?|Harv\.?\s*L\.?\s*Rev\.?|Stan\.?\s*L\.?\s*Rev\.?)\b',
        r'\b(?:Colum\.?\s*L\.?\s*Rev\.?|Mich\.?\s*L\.?\s*Rev\.?|Cornell\s+L\.?\s*Rev\.?)\b',
        r'\b(?:MULR|UNSWLJ|SydLR|UQLJ|UWALR|AdelLR|MonLR|MelbULawRw)\b',
        r'\b(?:AJLL|ABLR|AIAL\s+Forum|Fed(?:eral)?\s+L(?:aw)?\s+Rev(?:iew)?)\b',
        r'[A-Z][a-z]+,\s*"[^"]+"\s*\(\d{4}\)',
        r'\b(?:SAcLJ|Mal\.?\s*L\.?\s*R\.?|LMCLQ|JBL|ICLQ|AJCL|Sing\s+L\s+Rev)\b',
        r'"[^"]{10,}"\s*\(\d{4}\)\s+\d*\s*(?:SAcLJ|LQR|CLJ|OJLS|Sing\s+L\s+Rev|SJLS|LMCLQ|JBL|ICLQ)',
    ]

    book_patterns = [
        r'[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*,\s+[A-Z][A-Za-z\s:]+\([A-Za-z\s]+,\s*\d{4}\)',
        r'[A-Z][a-z]+,\s+[A-Z][A-Za-z\s]+\(\d+(?:st|nd|rd|th)\s+[Ee]d(?:ition)?,\s*\d{4}\)',
        r'\b(?:Halsbury|Chitty|Dicey|McGregor|Treitel|Anson|Cheshire|Winfield|Salmond)\b',
        r'\b(?:Snell|Bowstead|Phipson|Archbold|Lewin|Scrutton|Gatley|Keating)\b',
        r'\b(?:MacGillivray|Underhill|Spry|Bennion|Colinvaux|Williston|Corbin)\b',
        r'\b(?:Oppenheim|Brownlie|Pomeroy|Craies|Stroud|Odgers)\b',
        r'\b(?:Clerk\s*&\s*Lindsell|Goff\s*&\s*Jones|Spencer\s+Bower|Smith\s*&\s*Hogan)\b',
        r'\b(?:Mustill\s*&\s*Boyd|Megarry\s*&\s*Wade|Wade\s*&\s*Forsyth|Cross\s*&\s*Tapper)\b',
        r'\b(?:Bullen\s*&\s*Leake|Charlesworth\s*&\s*Percy|de\s+Smith)\b',
        r"\bBenjamin'?s?\s+(?:Sale|on\s+Sale)",
        r"\bFleming'?s?\s+(?:Law\s+of\s+Torts|Torts)",
        r"\bGower'?s?\s+(?:Principles|Company|Modern\s+Company)",
        r'\bSingapore\s+Civil\s+Procedure\b',
        r"\bMallal'?s?\s+Digest\b",
        r'\((?:Oxford\s+University\s+Press|Cambridge\s+University\s+Press|Hart\s+Publishing|Sweet\s*&\s*Maxwell|LexisNexis|Academy\s+Publishing|Butterworths|Thomson\s+Reuters|Clarendon\s+Press|Stevens|Law\s+Book\s+Co),\s*(?:\d+(?:st|nd|rd|th)\s+[Ee]d(?:ition)?,?\s*)?\d{4}\)',
        r'\([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?\s+University\s+Press,\s*(?:\d+(?:st|nd|rd|th)\s+[Ee]d(?:ition)?,?\s*)?\d{4}\)',
        r'\b(?:Ratanlal|Sarkar|Gour)\b(?:\s*&\s*(?:Dhirajlal|Thakore))?\S*\s+(?:Law\s+of|Indian|on\s+)',
    ]

    matches = []
    for p in journal_patterns:
        for m in re.finditer(p, text, re.IGNORECASE):
            matches.append((m.start(), m.end(), m.group()))
    for p in book_patterns:
        for m in re.finditer(p, text):
            matches.append((m.start(), m.end(), m.group()))

    # Sort by position, longest match first; remove overlaps
    matches.sort(key=lambda x: (x[0], -(x[1] - x[0])))
    used = []
    for s, e, t in matches:
        if any(s < ue and e > us for us, ue, _ in used):
            continue
        used.append((s, e, t))

    return len(used)


# ============================================================================
# MAIN PROCESSING
# ============================================================================

def process_file(filepath: Path, court_label: str) -> Optional[Dict]:
    """Process a single corpus file and return a metrics dict."""
    try:
        text = filepath.read_text(encoding='utf-8', errors='replace')
    except Exception as e:
        print(f'  ERROR reading {filepath.name}: {e}', flush=True)
        return None

    text = text.replace('\r\n', '\n').replace('\r', '\n')

    headnotes, core, footnotes = extract_sections(text)
    meta = extract_metadata(text, filepath.name)

    if not meta['court'] and court_label:
        meta['court'] = court_label

    # Citation & academic counts — raw core + footnotes (citations intact)
    cite_text = (core + '\n' + footnotes).strip()
    cites     = count_citations(cite_text)
    acad      = count_academic_references(cite_text)

    # Readability — core only, citations stripped
    metrics_text          = prepare_text_for_metrics(core)
    fkgl, fkre, smog, asl = compute_readability(metrics_text)

    return {
        'Case':            meta['case'],
        'Citation':        meta['citation'],
        'Year':            meta['year'],
        'Court':           meta['court'],
        'Date':            meta['date'],
        'Headnotes_Words': word_count(headnotes),
        'Core_Words':      word_count(core),
        'Footnote_Words':  word_count(footnotes),
        'FKGL':            fkgl,
        'FKRE':            fkre,
        'SMOG':            smog,
        'Avg_Sent_Length': asl,
        'Cites_UK':        cites['UK'],
        'Cites_AU':        cites['AU'],
        'Cites_SG':        cites['SG'],
        'Cites_NZ':        cites['NZ'],
        'Cites_CAN':       cites['CAN'],
        'Cites_USA':       cites['USA'],
        'Cites_EU':        cites['EU'],
        'Cites_HK':        cites['HK'],
        'Cites_MY':        cites['MY'],
        'Cites_IND':       cites['IND'],
        'Cites_OTHER':     cites['OTHER'],
        'Academic_Refs':   acad,
        'Filename':        filepath.name,
    }


def process_folder(input_folder: str, output_path: str, court_label: str = '') -> None:
    """Process all .txt corpus files in input_folder and write Excel output."""
    input_path = Path(input_folder)
    txt_files  = sorted(p for p in input_path.rglob('*.txt')
                        if not p.name.startswith('.') and not p.name.endswith('.bak'))

    if not txt_files:
        print(f'No .txt files found in {input_folder}', flush=True)
        return

    print(f'Found {len(txt_files)} files — processing...', flush=True)
    print('[INIT] Loading legal sentence tokeniser...', flush=True)
    get_legal_tokenizer()

    rows = []
    for i, fp in enumerate(txt_files, 1):
        row = process_file(fp, court_label)
        if row:
            rows.append(row)
        if i % 100 == 0:
            print(f'  {i}/{len(txt_files)} processed...', flush=True)

    print(f'Done — {len(rows)} rows.', flush=True)

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
        ws = writer.sheets['Data']

        # Header formatting
        hdr_fill = PatternFill(fill_type='solid', fgColor='1B2A6B')  # navy
        hdr_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        ws.freeze_panes = 'A2'

    print(f'Saved → {output_path}', flush=True)


# ============================================================================
# CLI ENTRY POINT
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Compute readability and citation metrics for cleaned judicial corpus files.'
    )
    parser.add_argument('--input',  required=True,
                        help='Folder containing cleaned .txt corpus files')
    parser.add_argument('--output', required=True,
                        help='Output Excel file path (.xlsx)')
    parser.add_argument('--court',  default='',
                        help='Court label to apply when not present in file header '
                             '(e.g. HCA, UKSC, SGCA)')
    args = parser.parse_args()
    process_folder(args.input, args.output, court_label=args.court)


if __name__ == '__main__':
    main()
